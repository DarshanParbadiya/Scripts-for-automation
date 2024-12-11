import pandas as pd
from openpyxl import load_workbook
def get_the_dealer_mapping_pld(file_paths):
    try:
        pld = pd.read_excel(file_paths.pld_path)
        pld.columns = pld.columns.str.replace('\n', ' ').str.strip()
        columns_to_keep_pld = [12,13]
        pld_dealer_codes = pld.iloc[:,columns_to_keep_pld]
        pld_dealer_codes.columns = ['DEALER NAME','DEALER CODE']
        
        ivra = pd.read_excel(file_paths.ivra_path)
        columns_to_keep_ivra = [34,35]
        ivra_dealer_codes = ivra.iloc[:,columns_to_keep_ivra]
        merged = pd.concat([pld_dealer_codes,ivra_dealer_codes],axis=0)
        # Drop duplicate rows based on the 'DEALER NAME' column
        unique_dealers = merged.drop_duplicates(subset='DEALER CODE')
        if file_paths.update_dealer_codes:
            # Read the specific sheet
            excel_file = pd.ExcelFile(file_paths.Transaction_file_path)
            sheet_name = file_paths.dealer_codes_sheet_name
            if sheet_name in excel_file.sheet_names:
                print(f"Sheet '{sheet_name}' exists in the Excel file.")
                # read the existing sheet

                existing_dealer_codes = pd.read_excel(file_paths.Transaction_file_path, sheet_name=sheet_name)
                # append newly created dealer codes with existing codes
                updated_dealer_codes = pd.concat([existing_dealer_codes,unique_dealers],axis=0)
                # Drop duplicate rows based on the 'DEALER CODE' column
                updated_dealer_codes = updated_dealer_codes.drop_duplicates(subset='DEALER CODE')
                
                clear_and_update_sheet(file_paths, sheet_name, updated_dealer_codes)
                
            else:
                print(f"Sheet '{sheet_name}' does not exist in the Excel file.")
                try:
                    with pd.ExcelWriter(file_paths.Transaction_file_path, engine='openpyxl', mode='a') as writer:
                        unique_dealers.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"Sheet '{sheet_name}' created in the Excel file.")
                except:
                    print('can not create the sheet in the file, so creating a csv file')
        
        return unique_dealers
    except Exception as e:
        print(f'can not create the dealer codes from PLD and IVRA files:',e)


def clear_and_update_sheet(file_paths, sheet_name, unique_dealers):
    try:
        # Load the workbook
        workbook = load_workbook(file_paths.Transaction_file_path)

        # Check if the sheet exists
        if sheet_name in workbook.sheetnames:
            # print(f"Sheet '{sheet_name}' exists in the Excel file.")
            
            # Access the sheet
            sheet = workbook[sheet_name]
            # delete_the_sheet(file_paths,sheet_name)
            
            # Clear the sheet content
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.value = None
            workbook.save(file_paths.Transaction_file_path)
            print(f"Cleared content of the sheet '{sheet_name}'.")
        
        # Write new content
        with pd.ExcelWriter(file_paths.Transaction_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            unique_dealers.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Sheet '{sheet_name}' updated with new content.")
    
    except Exception as e:
        print(f"An error occurred: {e}")
        print("Unable to update the Excel file. Saving the content as a CSV instead.")
        unique_dealers.to_csv(f"{sheet_name}.csv", index=False)
        print(f"CSV file '{sheet_name}.csv' created.")

def delete_the_sheet(file_paths,sheet_name):
    # Check if the sheet exists
    workbook = load_workbook(file_paths.Transaction_file_path)
    if sheet_name in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' exists in the Excel file.")
        
        # Remove the existing sheet
        del workbook[sheet_name]
        print(f"Sheet '{sheet_name}' removed.")
        workbook.save(file_paths.Transaction_file_path)
 

def map_dealer_codes(file_paths,df):
    codes = pd.read_excel(file_paths.Transaction_file_path, sheet_name=file_paths.dealer_codes_sheet_name)
    merged_df_pld = pd.merge(df, codes[['DEALER NAME', 'DEALER CODE']], on='DEALER CODE', how='left')
    return merged_df_pld